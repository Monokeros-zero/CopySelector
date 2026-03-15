#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Excel文件处理模块
"""

import pandas as pd
import os
import re
from openpyxl import load_workbook

class ExcelProcessor:
    """Excel文件处理类"""
    
    def __init__(self):
        """初始化Excel处理器"""
        self.month_aliases = self.load_month_aliases()
    
    def load_month_aliases(self):
        """加载月份别名配置"""
        month_json = os.path.join(os.path.dirname(__file__), "..", "month.json")
        if os.path.exists(month_json):
            try:
                import json
                with open(month_json, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except Exception as e:
                print(f"Error loading month.json: {e}")
        # 默认别名
        return {
            "1月": ["1月", "1月预测", "1月预估"],
            "2月": ["2月", "2月预测", "2月预估"],
            "3月": ["3月", "3月预测", "3月预估"],
            "4月": ["4月", "4月预测", "4月预估"],
            "5月": ["5月", "5月预测", "5月预估"],
            "6月": ["6月", "6月预测", "6月预估"],
            "7月": ["7月", "7月预测", "7月预估"],
            "8月": ["8月", "8月预测", "8月预估"],
            "9月": ["9月", "9月预测", "9月预估"],
            "10月": ["10月", "10月预测", "10月预估"],
            "11月": ["11月", "11月预测", "11月预估"],
            "12月": ["12月", "12月预测", "12月预估"]
        }
    
    def read_source_file(self, source_file, category_row, indicator_start_row):
        """读取参考文件，提取产品的勾选状态"""
        df = pd.read_excel(source_file, sheet_name='Sheet1')
        
        # 加载所有产品（从第1列开始）
        all_category_cols = df.columns[1:]
        
        # 提取产品名称
        category_row_idx = category_row - 1
        category_names = df.iloc[category_row_idx, 1:].tolist()
        
        # 提取指标行
        indicator_start_idx = indicator_start_row - 1
        indicators = df.iloc[indicator_start_idx:, 0].tolist()  # 第一列是指标名
        
        # 提取产品的勾选状态
        category_status = {}
        for i, col in enumerate(all_category_cols):
            if i < len(category_names):
                category = category_names[i]
                status = {}
                for j, indicator in enumerate(indicators):
                    cell_value = df.iloc[indicator_start_idx + j, df.columns.get_loc(col)]
                    status[indicator] = cell_value
                category_status[category] = status
        
        return category_status
    
    def read_target_file(self, target_file):
        """读取结果文件，获取sheet页信息"""
        xls = pd.ExcelFile(target_file)
        return xls.sheet_names
    
    def map_data(self, source_status, target_file, category_mapping, month_config, check_marks, target_indicator_start_row):
        """将参考文件的勾选状态映射到结果文件"""
        # 加载目标文件
        wb = load_workbook(target_file)
        
        # 处理每个产品
        for source_category, target_category in category_mapping.items():
            if source_category not in source_status:
                continue
            
            # 检查目标sheet是否存在
            if target_category not in wb.sheetnames:
                print(f"Sheet {target_category} not found in target file")
                continue
            
            # 获取目标sheet
            ws = wb[target_category]
            
            # 提取指标行
            indicator_start_row = target_indicator_start_row
            
            # 构建指标到行号的映射
            indicator_row_map = {}
            for row in range(indicator_start_row, ws.max_row + 1):
                indicator = ws.cell(row=row, column=1).value
                if indicator:
                    indicator_row_map[indicator] = row
            
            # 构建月份列到列号的映射（支持第1行和第2行）
            month_col_map = {}
            for row in [1, 2]:  # 检查第1行和第2行
                for col in range(1, ws.max_column + 1):
                    col_name = ws.cell(row=row, column=col).value
                    if col_name:
                        # 提取月份数字
                        match = re.search(r'\d+', str(col_name))
                        if match:
                            month = int(match.group())
                            # 月份可能在多个列出现，都需要标记
                            if month not in month_col_map:
                                month_col_map[month] = []
                            month_col_map[month].append(col)
                        else:
                            # 检查是否是月份别名
                            for month_str, aliases in self.month_aliases.items():
                                if str(col_name) in aliases:
                                    # 提取月份数字
                                    month_match = re.search(r'\d+', month_str)
                                    if month_match:
                                        month = int(month_match.group())
                                        if month not in month_col_map:
                                            month_col_map[month] = []
                                        month_col_map[month].append(col)
            
            # 先清空所有参考文件里的指标在目标文件中的对应指标行的月份列标记
            # 遍历所有参考文件中的指标
            for indicator, status in source_status[source_category].items():
                if indicator in indicator_row_map:
                    row = indicator_row_map[indicator]
                    # 清空所有月份列（包括月份别名）
                    for month, cols in month_col_map.items():
                        for col in cols:
                            ws.cell(row=row, column=col).value = None
            
            # 处理每个指标
            for indicator, status in source_status[source_category].items():
                if status in check_marks['source']:
                    if indicator in indicator_row_map:
                        row = indicator_row_map[indicator]
                        # 勾选的月份列
                        for month in month_config['selected_months']:
                            if month in month_col_map:
                                # 处理所有匹配的列
                                for col in month_col_map[month]:
                                    # 标记勾选，只修改标记，不改行头和列头
                                    ws.cell(row=row, column=col).value = check_marks['target']
                    else:
                        # 目标文件中没有的指标，打印出来但不打断处理
                        print(f"Indicator '{indicator}' not found in target sheet '{target_category}'")
            
            print(f"Updated sheet: {target_category}")
        
        # 保存文件
        try:
            wb.save(target_file)
        except Exception as e:
            if "Permission denied" in str(e) or "被另一进程使用" in str(e):
                raise Exception(f"目标文件无法写入，可能是文件被其他程序打开。请关闭文件后重试。")
            else:
                raise
